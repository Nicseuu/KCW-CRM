from __future__ import annotations

import hashlib
import io
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.excel import ExcelInventoryFile, ExcelInventoryRow
from app.models.inventory import Inventory
from app.models.product import Product

REQUIRED = ["SKU", "Product Name", "Total Stock"]


def sha256_bytes(b: bytes) -> str:
    return hashlib.sha256(b).hexdigest()


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_int_stock(value: Any) -> int:
    """
    Convert Excel cell to non-negative int stock.
    Accepts: int/float/str/None.
    """
    if value is None:
        return 0
    if isinstance(value, bool):
        return 0
    if isinstance(value, (int,)):
        return max(0, int(value))
    if isinstance(value, float):
        # Excel often stores numeric cells as float
        return max(0, int(value))
    s = str(value).strip()
    if not s:
        return 0
    try:
        # supports "10", "10.0"
        return max(0, int(float(s)))
    except Exception:
        return 0


def parse_excel_to_rows(file_bytes: bytes) -> list[dict]:
    """
    Expected columns (exact names, trimmed):
      - SKU
      - Product Name
      - Total Stock
    Reads the first sheet by default.
    """
    wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb.active

    # Read header row (row 1)
    header_cells = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    headers = [_normalize_header(h) for h in header_cells]

    # Build header index map
    idx = {h: i for i, h in enumerate(headers) if h}

    missing = [c for c in REQUIRED if c not in idx]
    if missing:
        raise ValueError(f"Missing columns: {missing}. Expected: {REQUIRED}")

    rows: list[dict] = []

    # Iterate data rows
    for row_cells in ws.iter_rows(min_row=2, values_only=True):
        sku = str(row_cells[idx["SKU"]] or "").strip()
        if not sku:
            continue

        product_name = str(row_cells[idx["Product Name"]] or "").strip()
        if not product_name:
            product_name = sku

        total_stock = _to_int_stock(row_cells[idx["Total Stock"]])

        raw = {
            "SKU": sku,
            "Product Name": product_name,
            "Total Stock": total_stock,
        }

        rows.append(
            {
                "sku": sku,
                "product_name": product_name,
                "total_stock": total_stock,
                "raw": raw,
            }
        )

    return rows


def create_excel_file_record(db: Session, *, org_id: int, filename: str, file_bytes: bytes) -> ExcelInventoryFile:
    h = sha256_bytes(file_bytes)
    excel_file = ExcelInventoryFile(org_id=org_id, filename=filename, content_hash=h, status="UPLOADED")
    db.add(excel_file)
    db.flush()
    return excel_file


def import_excel_file(db: Session, *, org_id: int, excel_file_id: int, file_bytes: bytes) -> dict:
    excel_file = db.get(ExcelInventoryFile, excel_file_id)
    if not excel_file or excel_file.org_id != org_id:
        raise ValueError("Excel file not found")

    rows = parse_excel_to_rows(file_bytes)

    # Save rows for audit
    db.query(ExcelInventoryRow).filter(ExcelInventoryRow.excel_file_id == excel_file_id).delete()
    for r in rows:
        db.add(
            ExcelInventoryRow(
                excel_file_id=excel_file_id,
                sku=r["sku"],
                product_name=r["product_name"],
                total_stock=r["total_stock"],
                raw=r["raw"],
            )
        )

    updated_inv = 0
    created_products = 0

    for r in rows:
        sku = r["sku"]
        name = r["product_name"]
        total = int(r["total_stock"])

        product = db.scalar(select(Product).where(Product.org_id == org_id, Product.sku == sku))
        if not product:
            product = Product(org_id=org_id, sku=sku, name=name)
            db.add(product)
            db.flush()
            created_products += 1
        else:
            if name and product.name != name:
                product.name = name

        inv = db.scalar(select(Inventory).where(Inventory.org_id == org_id, Inventory.product_id == product.id))
        if not inv:
            inv = Inventory(org_id=org_id, product_id=product.id, total_stock=0, reserved_stock=0, available_stock=0)
            db.add(inv)
            db.flush()

        inv.total_stock = total
        inv.available_stock = max(0, int(inv.total_stock) - int(inv.reserved_stock))
        updated_inv += 1

    excel_file.status = "PROCESSED"
    excel_file.error = None

    return {"rows": len(rows), "products_created": created_products, "inventory_updated": updated_inv}
